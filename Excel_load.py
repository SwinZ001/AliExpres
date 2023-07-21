# 引入速卖通类
from AliExpres import Ali_Express_Register
import openpyxl as openpyxl
import threading

class Excel_load():

    def __init__(self,ex):
        self.ex=ex
        xl = openpyxl.load_workbook(self.ex)
        sheets_name = xl.get_sheet_by_name('Sheet')
        # 获取总行数
        rows=sheets_name.max_row
        for i in range(1,rows+1):
            use=sheets_name.cell(row=i, column=1).value
            pwd=sheets_name.cell(row=i, column=2).value
            # 传入从表格读取的账号密码
            thread1=threading.Thread(name=i, target=Ali_Express_Register, args=(use, pwd))
            print(thread1.name)
            thread1.start()


if __name__ == "__main__":
    Exceld=Excel_load(r'E:\test.xlsx')